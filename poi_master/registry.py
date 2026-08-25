"""서울시 XLSX 기준정보와 Shapefile 영역을 하나의 검증된 POI Master로 결합한다."""

from __future__ import annotations

import hashlib
import io
import math
import re
import stat
import unicodedata
import warnings
import zipfile
from dataclasses import dataclass
from pathlib import PurePosixPath

import pyarrow as pa
import shapefile
import shapely
from core.poi_master import POI_MASTER_SCHEMA_VERSION
from openpyxl import load_workbook
from pyproj import CRS, Transformer
from shapely.geometry import shape
from shapely.geometry.base import BaseGeometry
from shapely.ops import transform as shapely_transform

from source import SourceAssets

POI_MASTER_SCHEMA = pa.schema(
    [
        ("AREA_CD", pa.string()),
        ("AREA_NM", pa.string()),
        ("CATEGORY", pa.string()),
        ("ENG_NM", pa.string()),
        ("SOURCE_NO", pa.int64()),
        ("GEOMETRY_WKB", pa.binary()),
        ("AREA_M2", pa.float64()),
    ]
)
"""Collector와 Normalizer가 공유하는 POI Master의 exact Arrow schema다."""

_AREA_CODE = re.compile(r"POI[0-9]{3}\Z")
_REQUIRED_XLSX_COLUMNS = ("CATEGORY", "NO", "AREA_CD", "AREA_NM", "ENG_NM")
_REQUIRED_SHAPE_EXTENSIONS = frozenset({".shp", ".shx", ".dbf", ".prj", ".cpg"})
_MAX_ZIP_MEMBERS = 100
_MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
_TO_EPSG5179 = Transformer.from_crs("EPSG:4326", "EPSG:5179", always_xy=True)
_SEOUL_WGS84_ENVELOPE = (126.5, 37.0, 127.5, 38.0)
_SEOUL_EPSG5179_ENVELOPE = (850_000.0, 1_850_000.0, 1_050_000.0, 2_050_000.0)


class PoiRegistryError(ValueError):
    """목록과 영역 자료가 POI Master 품질 계약을 위반했을 때 발생한다."""


@dataclass(frozen=True, slots=True)
class RegistryBuild:
    """게시할 Arrow table과 검증·원천 provenance를 함께 보관한다."""

    table: pa.Table
    repaired_count: int
    list_sha256: str
    areas_sha256: str


@dataclass(frozen=True, slots=True)
class _ListRow:
    """XLSX 한 행의 정규화된 기준정보를 표현한다."""

    area_cd: str
    area_nm: str
    category: str
    eng_nm: str
    source_no: int


@dataclass(frozen=True, slots=True)
class _ShapeRow:
    """Shapefile 한 행의 WGS84 geometry와 기준정보를 표현한다."""

    area_cd: str
    area_nm: str
    category: str
    geometry: BaseGeometry


def _required_text(value: object, field: str) -> str:
    """필수 셀 값을 NFC 문자열로 정규화하고 빈 값을 거부한다."""
    text = unicodedata.normalize("NFC", "" if value is None else str(value)).strip()
    if not text:
        raise PoiRegistryError(f"POI {field} 값이 비어 있습니다.")
    return text


def _source_number(value: object) -> int:
    """Excel의 NO 값을 손실 없는 양의 정수로 변환한다."""
    if isinstance(value, bool):
        raise PoiRegistryError(f"POI NO가 정수가 아닙니다: {value!r}")
    if isinstance(value, int):
        number = value
    elif isinstance(value, float) and value.is_integer():
        number = int(value)
    elif isinstance(value, str) and re.fullmatch(r"[0-9]+", value.strip()):
        number = int(value.strip())
    else:
        raise PoiRegistryError(f"POI NO가 정수가 아닙니다: {value!r}")
    if number < 1:
        raise PoiRegistryError(f"POI NO는 양수여야 합니다: {number}")
    return number


def _row_cell(
    raw_row: tuple[object, ...], positions: dict[str, int], name: str
) -> object:
    """짧은 XLSX 행에서 누락된 셀을 None으로 반환한다."""
    position = positions[name]
    return raw_row[position] if position < len(raw_row) else None


def _read_list_rows(payload: bytes) -> tuple[_ListRow, ...]:
    """XLSX 첫 worksheet를 읽어 exact 필드의 고유 POI 목록으로 변환한다."""
    try:
        # 공식 파일의 data-validation extension은 기준정보 셀과 무관하고 read-only
        # openpyxl이 지원하지 않는다는 경고만 낸다. 다른 경고는 숨기지 않는다.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Unknown extension is not supported and will be removed",
                category=UserWarning,
                module=r"openpyxl\.worksheet\._reader",
            )
            workbook = load_workbook(
                io.BytesIO(payload), read_only=True, data_only=True, keep_links=False
            )
            worksheet = workbook.worksheets[0]
            all_rows = list(worksheet.iter_rows(values_only=True))
        raw_header = all_rows[0]
        rows = iter(all_rows[1:])
    except (
        OSError,
        ValueError,
        KeyError,
        IndexError,
        StopIteration,
        zipfile.BadZipFile,
    ) as exc:
        raise PoiRegistryError("서울시 POI 목록 XLSX를 읽을 수 없습니다.") from exc
    try:
        header_values = list(raw_header)
        while header_values and header_values[-1] is None:
            header_values.pop()
        header = tuple(_required_text(value, "XLSX header") for value in header_values)
        if len(header) != len(set(header)):
            raise PoiRegistryError("POI 목록 XLSX header에 중복 컬럼이 있습니다.")
        missing = [column for column in _REQUIRED_XLSX_COLUMNS if column not in header]
        if missing:
            raise PoiRegistryError(
                f"POI 목록 XLSX에 필수 컬럼이 없습니다: missing={missing}"
            )
        positions = {name: header.index(name) for name in _REQUIRED_XLSX_COLUMNS}
        parsed: list[_ListRow] = []
        for raw_row in rows:
            if not any(value is not None and str(value).strip() for value in raw_row):
                continue

            area_cd = _required_text(
                _row_cell(raw_row, positions, "AREA_CD"), "AREA_CD"
            )
            if _AREA_CODE.fullmatch(area_cd) is None:
                raise PoiRegistryError(f"POI AREA_CD 형식이 잘못됐습니다: {area_cd!r}")
            parsed.append(
                _ListRow(
                    area_cd=area_cd,
                    area_nm=_required_text(
                        _row_cell(raw_row, positions, "AREA_NM"), "AREA_NM"
                    ),
                    category=_required_text(
                        _row_cell(raw_row, positions, "CATEGORY"), "CATEGORY"
                    ),
                    eng_nm=_required_text(
                        _row_cell(raw_row, positions, "ENG_NM"), "ENG_NM"
                    ),
                    source_no=_source_number(_row_cell(raw_row, positions, "NO")),
                )
            )
    finally:
        workbook.close()

    if not parsed:
        raise PoiRegistryError("POI 목록 XLSX에 데이터 행이 없습니다.")
    codes = [row.area_cd for row in parsed]
    numbers = [row.source_no for row in parsed]
    if len(codes) != len(set(codes)):
        raise PoiRegistryError("POI 목록 XLSX에 중복 AREA_CD가 있습니다.")
    if len(numbers) != len(set(numbers)):
        raise PoiRegistryError("POI 목록 XLSX에 중복 NO가 있습니다.")
    return tuple(parsed)


def _safe_shape_members(payload: bytes) -> dict[str, bytes]:
    """영역 ZIP을 추출하지 않고 안전성과 한 shapefile 묶음을 검증해 읽는다."""
    try:
        archive = zipfile.ZipFile(io.BytesIO(payload))
    except (OSError, zipfile.BadZipFile) as exc:
        raise PoiRegistryError("서울시 POI 영역 ZIP을 열 수 없습니다.") from exc
    with archive:
        infos = archive.infolist()
        if not infos or len(infos) > _MAX_ZIP_MEMBERS:
            raise PoiRegistryError(
                f"POI 영역 ZIP member 수가 안전 범위를 벗어났습니다: {len(infos)}"
            )
        if sum(info.file_size for info in infos) > _MAX_UNCOMPRESSED_BYTES:
            raise PoiRegistryError("POI 영역 ZIP의 압축 해제 크기가 제한을 넘습니다.")
        members: dict[str, bytes] = {}
        stems: set[str] = set()
        for info in infos:
            normalized = info.filename.replace("\\", "/")
            path = PurePosixPath(normalized)
            file_mode = (info.external_attr >> 16) & 0o170000
            if (
                normalized.startswith("/")
                or re.match(r"[A-Za-z]:", normalized)
                or ".." in path.parts
                or file_mode == stat.S_IFLNK
            ):
                raise PoiRegistryError(
                    f"POI 영역 ZIP에 안전하지 않은 경로가 있습니다: {info.filename!r}"
                )
            if info.is_dir():
                continue
            suffix = path.suffix.casefold()
            if suffix not in _REQUIRED_SHAPE_EXTENSIONS:
                continue
            stem = path.stem.casefold()
            stems.add(stem)
            if suffix in members:
                raise PoiRegistryError(f"POI 영역 ZIP에 {suffix} 파일이 중복됐습니다.")
            members[suffix] = archive.read(info)
        if archive.testzip() is not None:
            raise PoiRegistryError("POI 영역 ZIP CRC 검증에 실패했습니다.")
    missing = sorted(_REQUIRED_SHAPE_EXTENSIONS - members.keys())
    if missing:
        raise PoiRegistryError(f"POI 영역 ZIP sidecar가 없습니다: missing={missing}")
    if len(stems) != 1:
        raise PoiRegistryError(
            f"POI 영역 ZIP sidecar의 파일 stem이 다릅니다: stems={sorted(stems)}"
        )
    return members


def _validate_shape_crs(members: dict[str, bytes]) -> None:
    """PRJ와 CPG가 WGS84·UTF-8 원천 계약을 나타내는지 확인한다."""
    try:
        crs = CRS.from_wkt(members[".prj"].decode("utf-8-sig").strip())
    except (UnicodeDecodeError, ValueError) as exc:
        raise PoiRegistryError("POI 영역 PRJ를 해석할 수 없습니다.") from exc
    wgs84 = CRS.from_epsg(4326)
    if crs.to_epsg() != 4326 and not crs.equals(wgs84, ignore_axis_order=True):
        raise PoiRegistryError(f"POI 영역 CRS가 WGS84가 아닙니다: {crs.to_string()}")
    try:
        encoding = members[".cpg"].decode("ascii").strip().upper().replace("_", "-")
    except UnicodeDecodeError as exc:
        raise PoiRegistryError("POI 영역 CPG를 해석할 수 없습니다.") from exc
    if encoding not in {"UTF-8", "UTF8", "65001"}:
        raise PoiRegistryError(
            f"POI 영역 DBF encoding이 UTF-8이 아닙니다: {encoding!r}"
        )


def _read_shape_rows(payload: bytes) -> tuple[_ShapeRow, ...]:
    """영역 ZIP의 shapefile을 메모리에서 읽어 WGS84 geometry 행으로 변환한다."""
    members = _safe_shape_members(payload)
    _validate_shape_crs(members)
    try:
        reader = shapefile.Reader(
            shp=io.BytesIO(members[".shp"]),
            shx=io.BytesIO(members[".shx"]),
            dbf=io.BytesIO(members[".dbf"]),
            encoding="utf-8",
        )
        fields = {field[0] for field in reader.fields[1:]}
        missing_fields = {"AREA_CD", "AREA_NM", "CATEGORY"} - fields
        if missing_fields:
            raise PoiRegistryError(
                f"POI 영역 DBF에 필수 필드가 없습니다: missing={sorted(missing_fields)}"
            )
        rows: list[_ShapeRow] = []
        for shape_record in reader.iterShapeRecords():
            record = shape_record.record.as_dict()
            area_cd = _required_text(record.get("AREA_CD"), "shape AREA_CD")
            if _AREA_CODE.fullmatch(area_cd) is None:
                raise PoiRegistryError(
                    f"POI shape AREA_CD 형식이 잘못됐습니다: {area_cd!r}"
                )
            rows.append(
                _ShapeRow(
                    area_cd=area_cd,
                    area_nm=_required_text(record.get("AREA_NM"), "shape AREA_NM"),
                    category=_required_text(record.get("CATEGORY"), "shape CATEGORY"),
                    geometry=shape(shape_record.shape.__geo_interface__),
                )
            )
        reader.close()
    except PoiRegistryError:
        raise
    except (OSError, ValueError, shapefile.ShapefileException) as exc:
        raise PoiRegistryError("POI 영역 Shapefile을 읽을 수 없습니다.") from exc
    if not rows:
        raise PoiRegistryError("POI 영역 Shapefile에 데이터 행이 없습니다.")
    codes = [row.area_cd for row in rows]
    if len(codes) != len(set(codes)):
        raise PoiRegistryError("POI 영역 Shapefile에 중복 AREA_CD가 있습니다.")
    return tuple(rows)


def _largest_polygon(geometry: BaseGeometry, area_cd: str) -> BaseGeometry:
    """복구 결과에서 면적이 가장 큰 단일 Polygon을 선택한다."""
    if geometry.geom_type == "Polygon":
        return geometry
    if geometry.geom_type in {"MultiPolygon", "GeometryCollection"}:
        polygons = [part for part in geometry.geoms if part.geom_type == "Polygon"]
        if polygons:
            return max(polygons, key=lambda polygon: polygon.area)
    raise PoiRegistryError(
        f"{area_cd}: Polygon으로 만들 수 없는 geometry입니다: {geometry.geom_type}"
    )


def _validated_geometry(
    geometry_wgs84: BaseGeometry, area_cd: str
) -> tuple[BaseGeometry, bool]:
    """원천 geometry를 복구하고 EPSG:5179의 유효한 단일 Polygon으로 변환한다."""
    if geometry_wgs84.is_empty:
        raise PoiRegistryError(f"{area_cd}: geometry가 비어 있습니다.")
    if not _bounds_within(geometry_wgs84.bounds, _SEOUL_WGS84_ENVELOPE):
        raise PoiRegistryError(
            f"{area_cd}: WGS84 geometry가 서울 안전 범위를 벗어났습니다: "
            f"bounds={geometry_wgs84.bounds}"
        )
    repaired = not geometry_wgs84.is_valid or geometry_wgs84.geom_type != "Polygon"
    candidate = (
        shapely.make_valid(geometry_wgs84)
        if not geometry_wgs84.is_valid
        else geometry_wgs84
    )
    polygon = _largest_polygon(candidate, area_cd)
    projected = shapely_transform(_TO_EPSG5179.transform, polygon)
    if (
        projected.is_empty
        or not projected.is_valid
        or projected.geom_type != "Polygon"
        or projected.area <= 0
        or not _bounds_within(projected.bounds, _SEOUL_EPSG5179_ENVELOPE)
    ):
        raise PoiRegistryError(f"{area_cd}: EPSG:5179 geometry 검증에 실패했습니다.")
    return projected, repaired


def _bounds_within(
    bounds: tuple[float, float, float, float],
    envelope: tuple[float, float, float, float],
) -> bool:
    """Geometry bounds가 유한하고 안전 envelope 안에 완전히 포함되는지 반환한다."""
    min_x, min_y, max_x, max_y = bounds
    envelope_min_x, envelope_min_y, envelope_max_x, envelope_max_y = envelope
    return (
        all(math.isfinite(value) for value in bounds)
        and min_x < max_x
        and min_y < max_y
        and envelope_min_x <= min_x
        and max_x <= envelope_max_x
        and envelope_min_y <= min_y
        and max_y <= envelope_max_y
    )


def _metadata(
    assets: SourceAssets, list_sha256: str, areas_sha256: str
) -> dict[bytes, bytes]:
    """Parquet schema에 결합할 원천 provenance metadata를 만든다."""
    values = {
        "poi_master_schema_version": POI_MASTER_SCHEMA_VERSION,
        "geometry_crs": "EPSG:5179",
        "source_page_url": assets.page_url,
        "list_filename": assets.list_attachment.filename,
        "list_sequence": assets.list_attachment.sequence,
        "list_modified_date": assets.list_attachment.modified_date,
        "list_declared_place_count": str(
            assets.list_attachment.declared_place_count
        ),
        "list_sha256": list_sha256,
        "areas_filename": assets.areas_attachment.filename,
        "areas_sequence": assets.areas_attachment.sequence,
        "areas_modified_date": assets.areas_attachment.modified_date,
        "areas_declared_place_count": str(
            assets.areas_attachment.declared_place_count
        ),
        "areas_sha256": areas_sha256,
    }
    return {key.encode("utf-8"): value.encode("utf-8") for key, value in values.items()}


def build_registry(assets: SourceAssets) -> RegistryBuild:
    """두 공식 첨부를 교차 검증해 정렬된 EPSG:5179 POI Master를 만든다."""
    list_sha256 = hashlib.sha256(assets.list_bytes).hexdigest()
    areas_sha256 = hashlib.sha256(assets.areas_bytes).hexdigest()
    list_rows = _read_list_rows(assets.list_bytes)
    shape_rows = _read_shape_rows(assets.areas_bytes)
    list_by_code = {row.area_cd: row for row in list_rows}
    shape_by_code = {row.area_cd: row for row in shape_rows}
    if list_by_code.keys() != shape_by_code.keys():
        raise PoiRegistryError(
            "POI 목록과 영역의 AREA_CD 집합이 다릅니다: "
            f"list_only={sorted(list_by_code.keys() - shape_by_code.keys())}, "
            f"areas_only={sorted(shape_by_code.keys() - list_by_code.keys())}"
        )

    output_rows: list[dict[str, object]] = []
    repaired_count = 0
    for area_cd in sorted(list_by_code):
        list_row = list_by_code[area_cd]
        shape_row = shape_by_code[area_cd]
        if list_row.area_nm != shape_row.area_nm:
            raise PoiRegistryError(
                f"{area_cd}: 목록과 영역의 AREA_NM이 다릅니다: "
                f"list={list_row.area_nm!r}, areas={shape_row.area_nm!r}"
            )
        if list_row.category != shape_row.category:
            raise PoiRegistryError(
                f"{area_cd}: 목록과 영역의 CATEGORY가 다릅니다: "
                f"list={list_row.category!r}, areas={shape_row.category!r}"
            )
        projected, repaired = _validated_geometry(shape_row.geometry, area_cd)
        repaired_count += int(repaired)
        output_rows.append(
            {
                "AREA_CD": area_cd,
                "AREA_NM": list_row.area_nm,
                "CATEGORY": list_row.category,
                "ENG_NM": list_row.eng_nm,
                "SOURCE_NO": list_row.source_no,
                "GEOMETRY_WKB": shapely.to_wkb(
                    projected, byte_order=1, output_dimension=2, include_srid=False
                ),
                "AREA_M2": float(projected.area),
            }
        )
    table = pa.Table.from_pylist(output_rows, schema=POI_MASTER_SCHEMA)
    place_counts = {
        "list_filename": assets.list_attachment.declared_place_count,
        "areas_filename": assets.areas_attachment.declared_place_count,
        "xlsx_unique_area_cd": len(list_rows),
        "shape_unique_area_cd": len(shape_rows),
        "master_rows": table.num_rows,
    }
    if len(set(place_counts.values())) != 1:
        raise PoiRegistryError(
            "POI 첨부 선언과 실제 장소 수가 일치하지 않습니다: "
            f"counts={place_counts}"
        )
    table = table.replace_schema_metadata(_metadata(assets, list_sha256, areas_sha256))
    return RegistryBuild(
        table=table,
        repaired_count=repaired_count,
        list_sha256=list_sha256,
        areas_sha256=areas_sha256,
    )
