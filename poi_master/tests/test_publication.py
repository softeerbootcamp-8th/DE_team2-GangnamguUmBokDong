"""POI Master의 불변 S3 게시와 실패 시 활성본 보존을 검증한다."""

import io
from dataclasses import replace
from datetime import UTC, datetime, timedelta

import boto3
import pytest
from core.gold_publication.errors import ObjectCollisionError, ObjectStoreAccessError
from core.gold_publication.storage import ImmutablePutOutcome, S3ImmutableObjectStore
from core.poi_master import read_poi_master, resolve_poi_master
from moto import mock_aws
from openpyxl import load_workbook

import publication
from registry import PoiRegistryError
from tests.conftest import real_source_assets

BUCKET = "poi-master-test"
FIRST = datetime(2026, 8, 25, 1, 0, tzinfo=UTC)


@pytest.fixture(autouse=True)
def aws(monkeypatch):
    """각 테스트에 빈 mock S3 bucket과 고정 credential을 제공한다."""
    monkeypatch.setenv("AWS_ACCESS_KEY_ID", "testing")
    monkeypatch.setenv("AWS_SECRET_ACCESS_KEY", "testing")
    monkeypatch.setenv("AWS_DEFAULT_REGION", "ap-northeast-2")
    monkeypatch.setenv("S3_BUCKET", BUCKET)
    monkeypatch.delenv("S3_ENDPOINT_URL", raising=False)
    with mock_aws():
        boto3.client("s3", region_name="ap-northeast-2").create_bucket(
            Bucket=BUCKET,
            CreateBucketConfiguration={"LocationConstraint": "ap-northeast-2"},
        )
        yield


def _keys(prefix: str = "") -> list[str]:
    """mock bucket의 key를 정렬해 반환한다."""
    response = boto3.client("s3", region_name="ap-northeast-2").list_objects_v2(
        Bucket=BUCKET, Prefix=prefix
    )
    return sorted(item["Key"] for item in response.get("Contents", []))


def test_first_refresh_publishes_raw_silver_manifest_then_activation(
    monkeypatch,
) -> None:
    """최초 정상본이 모든 불변 계층과 활성 ref를 만든다."""
    writes: list[tuple[str, bool]] = []

    class RecordingStore(S3ImmutableObjectStore):
        """Producer가 요청한 모든 immutable put을 기록한다."""

        def put_once(
            self,
            uri: str,
            payload: bytes,
            *,
            expected_sha256: str | None = None,
            require_canonical_json: bool = False,
        ) -> ImmutablePutOutcome:
            """호출 계약을 기록하고 실제 조건부 S3 PUT을 실행한다."""
            writes.append((uri, require_canonical_json))
            return super().put_once(
                uri,
                payload,
                expected_sha256=expected_sha256,
                require_canonical_json=require_canonical_json,
            )

    monkeypatch.setattr(publication, "S3ImmutableObjectStore", RecordingStore)
    result = publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)

    assert result.status == "published"
    assert result.row_count == 121
    assert result.ref.mode == "s3"
    assert len(_keys("source_snapshot_raw/poi_master/")) == 2
    assert _keys("bronze/poi_master/") == []
    assert len(_keys("silver/poi_master/")) == 1
    assert len(_keys("source_snapshot_manifest/poi_master/")) == 1
    assert len(_keys("source_snapshot_pointer/poi_master/")) == 1
    assert resolve_poi_master(FIRST) == result.ref
    metadata = read_poi_master(result.ref).schema.metadata or {}
    assert metadata[b"list_uri"].startswith(
        f"s3://{BUCKET}/source_snapshot_raw/poi_master/list/".encode()
    )
    assert metadata[b"areas_uri"].startswith(
        f"s3://{BUCKET}/source_snapshot_raw/poi_master/areas/".encode()
    )
    assert len(writes) == 4
    assert [canonical for _uri, canonical in writes] == [False, False, False, True]
    write_prefixes = [
        uri.split(f"s3://{BUCKET}/", 1)[1].split("/", 1)[0]
        for uri, _canonical in writes
    ]
    assert write_prefixes == [
        "source_snapshot_raw",
        "source_snapshot_raw",
        "silver",
        "source_snapshot_manifest",
    ]


def test_same_content_next_day_is_unchanged_without_new_release() -> None:
    """HTML 날짜나 실행일이 아니라 두 첨부의 content hash가 같으면 재게시하지 않는다."""
    first = publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)
    before = _keys()

    second = publication.refresh_poi_master(
        real_source_assets(), activated_at=FIRST + timedelta(days=1)
    )

    assert second.status == "unchanged"
    assert second.ref == first.ref
    assert _keys() == before


def test_changed_filename_count_with_same_bytes_is_validated() -> None:
    """파일 bytes가 같아도 새 파일명 선언 수가 다르면 unchanged로 우회하지 않는다."""
    publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)
    source = real_source_assets()
    changed = replace(
        source,
        list_attachment=replace(
            source.list_attachment,
            filename="서울시 주요 125장소 목록.xlsx",
            declared_place_count=125,
        ),
        areas_attachment=replace(
            source.areas_attachment,
            filename="서울시 주요 125장소 영역.zip",
            declared_place_count=125,
        ),
    )

    with pytest.raises(PoiRegistryError, match="실제 장소 수"):
        publication.refresh_poi_master(
            changed,
            activated_at=FIRST + timedelta(days=1),
        )


def test_same_source_is_republished_when_schema_version_is_stale(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """원천 checksum이 같아도 이전 schema version이면 현재 계약으로 다시 게시한다."""
    first = publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)
    current_snapshot = publication._read_refresh_snapshot(first.ref)
    assert current_snapshot.table is not None
    stale_metadata = dict(current_snapshot.table.schema.metadata or {})
    stale_metadata[b"poi_master_schema_version"] = b"poi-master-v0"
    stale_snapshot = replace(
        current_snapshot,
        table=current_snapshot.table.replace_schema_metadata(stale_metadata),
    )
    monkeypatch.setattr(
        publication,
        "_read_refresh_snapshot",
        lambda _ref: stale_snapshot,
    )

    second = publication.refresh_poi_master(
        real_source_assets(),
        activated_at=FIRST + timedelta(days=1),
    )

    assert second.status == "published"
    assert second.ref != first.ref
    assert read_poi_master(second.ref).schema.metadata[
        b"poi_master_schema_version"
    ] == publication.POI_MASTER_SCHEMA_VERSION.encode("utf-8")


def test_stale_schema_cannot_weaken_legacy_drop_guard(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """행이 적은 알 수 없는 구버전은 repository 정상본보다 약한 기준이 될 수 없다."""
    first = publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)
    current_snapshot = publication._read_refresh_snapshot(first.ref)
    assert current_snapshot.table is not None
    stale_metadata = dict(current_snapshot.table.schema.metadata or {})
    stale_metadata[b"poi_master_schema_version"] = b"poi-master-v0"
    stale_snapshot = replace(
        current_snapshot,
        table=current_snapshot.table.slice(0, 1).replace_schema_metadata(
            stale_metadata
        ),
    )
    complete = publication.build_registry(real_source_assets())
    reduced = replace(complete, table=complete.table.slice(0, 96))
    monkeypatch.setattr(
        publication,
        "_read_refresh_snapshot",
        lambda _ref: stale_snapshot,
    )
    monkeypatch.setattr(publication, "build_registry", lambda _assets: reduced)

    with pytest.raises(
        publication.PoiPublicationError,
        match=r"previous=121, candidate=96",
    ):
        publication.refresh_poi_master(
            real_source_assets(),
            activated_at=FIRST + timedelta(days=1),
        )


def test_stale_schema_rejects_untrusted_area_code_identity(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """알 수 없는 schema의 중복 POI 행 수는 migration 감소 기준으로 신뢰하지 않는다."""
    first = publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)
    current_snapshot = publication._read_refresh_snapshot(first.ref)
    assert current_snapshot.table is not None
    stale_metadata = dict(current_snapshot.table.schema.metadata or {})
    stale_metadata[b"poi_master_schema_version"] = b"poi-master-v0"
    duplicated = current_snapshot.table.take([0, 0]).replace_schema_metadata(
        stale_metadata
    )
    stale_snapshot = replace(current_snapshot, table=duplicated)
    monkeypatch.setattr(
        publication,
        "_read_refresh_snapshot",
        lambda _ref: stale_snapshot,
    )

    with pytest.raises(publication.PoiPublicationError, match="AREA_CD identity"):
        publication.refresh_poi_master(
            real_source_assets(),
            activated_at=FIRST + timedelta(days=1),
        )


def test_stale_manifest_config_republishes_current_schema(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Table metadata와 원천 hash가 같아도 stale config version은 재게시한다."""
    first = publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)
    current_snapshot = publication._read_refresh_snapshot(first.ref)
    stale_snapshot = replace(
        current_snapshot,
        manifest=replace(
            current_snapshot.manifest,
            config_version=f"sha256:{'0' * 64}",
        ),
    )
    monkeypatch.setattr(
        publication,
        "_read_refresh_snapshot",
        lambda _ref: stale_snapshot,
    )

    second = publication.refresh_poi_master(
        real_source_assets(),
        activated_at=FIRST + timedelta(days=1),
    )

    assert second.status == "published"
    assert second.ref != first.ref


def test_activation_failure_leaves_previous_ref_selected(monkeypatch) -> None:
    """마지막 activation 쓰기가 실패하면 기존 활성본 선택은 바뀌지 않는다."""
    first = publication.refresh_poi_master(real_source_assets(), activated_at=FIRST)
    original = publication.activate_poi_master

    def fail_activation(**_kwargs):
        """S3 activation 단계의 장애를 흉내 낸다."""
        raise publication.PoiPublicationError("activation failure")

    monkeypatch.setattr(publication, "activate_poi_master", fail_activation)
    source = real_source_assets()
    workbook = load_workbook(io.BytesIO(source.list_bytes))
    worksheet = workbook.worksheets[0]
    eng_column = next(cell.column for cell in worksheet[1] if cell.value == "ENG_NM")
    worksheet.cell(row=2, column=eng_column).value = "Changed English Name"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    changed = type(source)(
        list_attachment=source.list_attachment,
        areas_attachment=source.areas_attachment,
        list_bytes=output.getvalue(),
        areas_bytes=source.areas_bytes,
    )
    with pytest.raises(publication.PoiPublicationError, match="activation failure"):
        publication.refresh_poi_master(changed, activated_at=FIRST + timedelta(days=1))

    monkeypatch.setattr(publication, "activate_poi_master", original)
    assert resolve_poi_master(FIRST + timedelta(days=1)) == first.ref


def test_first_refresh_rejects_drop_against_legacy_shapefile(monkeypatch) -> None:
    """활성본이 없어도 기존 Shapefile 121행 대비 20% 초과 감소를 차단한다."""
    source = real_source_assets()
    complete = publication.build_registry(source)
    reduced = replace(complete, table=complete.table.slice(0, 96))
    monkeypatch.setattr(publication, "build_registry", lambda _assets: reduced)

    assert publication._legacy_poi_count() == 121
    with pytest.raises(
        publication.PoiPublicationError,
        match=r"previous=121, candidate=96",
    ):
        publication.refresh_poi_master(source, activated_at=FIRST)

    assert _keys() == []


def test_immutable_write_collision_is_wrapped_without_overwrite() -> None:
    """조건부 PUT 충돌은 publication 오류로 변환하고 기존 bytes를 보존한다."""
    key = "source_snapshot_raw/poi_master/list/sha256=fixture.xlsx"
    client = boto3.client("s3", region_name="ap-northeast-2")
    client.put_object(Bucket=BUCKET, Key=key, Body=b"existing")

    with pytest.raises(publication.PoiPublicationError) as caught:
        publication._put_once(S3ImmutableObjectStore(), key, b"incoming")

    assert isinstance(caught.value.__cause__, ObjectCollisionError)
    persisted = client.get_object(Bucket=BUCKET, Key=key)["Body"].read()
    assert persisted == b"existing"


def test_immutable_write_reconciles_applied_put_response_loss() -> None:
    """PUT 반영 뒤 응답만 유실되면 exact readback으로 성공을 확인한다."""
    real_store = S3ImmutableObjectStore()

    class AppliedThenFailedStore(S3ImmutableObjectStore):
        """실제 객체를 쓴 뒤 transport 실패를 흉내 내는 store다."""

        def put_once(self, *args, **kwargs):
            """조건부 PUT을 완료한 뒤 응답 유실 오류를 발생시킨다."""
            real_store.put_once(*args, **kwargs)
            raise ObjectStoreAccessError("응답 유실")

        def read_bytes(self, *args, **kwargs):
            """실제 store에서 incoming checksum의 exact bytes를 읽는다."""
            return real_store.read_bytes(*args, **kwargs)

    key = "source_snapshot_raw/poi_master/list/sha256=response-lost.xlsx"
    publication._put_once(AppliedThenFailedStore(), key, b"persisted")

    payload = (
        boto3.client("s3", region_name="ap-northeast-2")
        .get_object(
            Bucket=BUCKET,
            Key=key,
        )["Body"]
        .read()
    )
    assert payload == b"persisted"
